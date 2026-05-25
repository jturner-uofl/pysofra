"""Tests for the Excel (.xlsx) renderer."""

from __future__ import annotations

import pytest

import pysofra as ps


class TestXlsx:
    def test_writes_file(self, small_trial, tmp_path):
        pytest.importorskip("xlsxwriter")
        out = tmp_path / "demo.xlsx"
        ps.tbl_one(small_trial, by="arm").add_p().to_xlsx(out)
        assert out.exists()
        assert out.stat().st_size > 2000

    def test_zip_signature(self, small_trial, tmp_path):
        # .xlsx is a zip; first 2 bytes should be "PK".
        pytest.importorskip("xlsxwriter")
        out = tmp_path / "demo.xlsx"
        ps.tbl_one(small_trial, by="arm").to_xlsx(out)
        with open(out, "rb") as f:
            magic = f.read(2)
        assert magic == b"PK"

    def test_with_caption_and_footnote(self, small_trial, tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        out = tmp_path / "demo.xlsx"
        (
            ps.tbl_one(small_trial, by="arm")
              .add_p()
              .set_caption("Baseline characteristics")
              .add_footnote("Synthetic data.")
              .to_xlsx(out)
        )
        wb = load_workbook(out)
        ws = wb.active
        all_text = " ".join(
            str(cell.value) for row in ws.iter_rows() for cell in row
            if cell.value is not None
        )
        assert "Baseline characteristics" in all_text
        assert "Synthetic data." in all_text

    def test_custom_sheet_name(self, small_trial, tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        out = tmp_path / "demo.xlsx"
        ps.tbl_one(small_trial, by="arm").to_xlsx(out, sheet_name="MyTable")
        wb = load_workbook(out)
        assert "MyTable" in wb.sheetnames
